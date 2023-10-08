import re
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from language_type import *

# android翻译源文件文件夹
android_trans_file_path = './trans_to_excel/input/android/'

# ios
ios_trans_file_path = './trans_to_excel/input/ios/'

# web翻译源文件文件夹
web_trans_file_path = './trans_to_excel/input/web/'

# 输出安卓excel文件
output_android_excel_path = './trans_to_excel/output/android/android_trans.xlsx'

# 输出iosexcel文件
output_ios_excel_path = './trans_to_excel/output/ios/ios_trans.xlsx'

# 输出webexcel文件
output_web_excel_path = './trans_to_excel/output/web/web_trans.xlsx'

# 自定义字体样式
font = Font(
    name="微软雅黑",  # 字体
    size=13,         # 字体大小
    color="424357",  # 字体颜色，用16进制rgb表示
    bold=False,       # 是否加粗，True/False
    italic=False,     # 是否斜体，True/False
    strike=None,     # 是否使用删除线，True/False
    underline=None,  # 下划线, 可选'singleAccounting', 'double', 'single', 'doubleAccounting'
)
align = Alignment(
    horizontal='left',     # 水平对齐，可选general、left、center、right、fill、justify、centerContinuous、distributed
    vertical='center',        # 垂直对齐， 可选top、center、bottom、justify、distributed
    text_rotation=0,       # 字体旋转，0~180整数
    wrap_text=True,       # 是否自动换行
    shrink_to_fit=False,   # 是否缩小字体填充
    indent=0,              # 缩进值
)
excel_cell_width = 80
excel_cell_height = 30
        
# 逐行读取ios国际化源文件
def read_trans_file_ios():
    # 多语言翻译的map 例如:{zh_hans: {'trans0001': "中文"}, en: {'trans0001': "english"}}
    trans_map = {}
    dir_path = f'{ios_trans_file_path}'
    file_list = os.listdir(dir_path)
    for file_name in file_list:
        if ".lproj" in file_name:
            trans_lang_map = _read_string_file(file_name)
            _key = file_name.replace('.lproj', '')
            trans_map[_key] = trans_lang_map
    # print('trans_map: ', trans_map)
    return trans_map

#  处理key值
def _read_string_file(file_name):
    file_path = f'{ios_trans_file_path}{file_name}/Localizable.strings'
    with open(file_path) as file:
        context = file.read()
        lines = context.split(';')
        # 单个文件翻译的map 例如:{'trans0001': "中文"}
        trans_lang_map = {}
        for line in lines:
            trans = line.split('=')
            if len(trans) > 1:
                # key 去掉空格 大写转小写
                key = trans[0].replace(' ', '').replace('\n', '').replace('\t', '').lower()
                text = _convert_to_excel_args(trans[1], PlatformType.ios)
                value = text.replace('None', '').replace('"', '')
                trans_lang_map[key] = value
        return trans_lang_map
    
    # 将iOS 安卓 web各端参数转换成excel统一格式{1}  eg: $%1$@起投 =》 ${1}起投
def _convert_to_excel_args(text, platform):
    value = text
    if platform == PlatformType.ios:
        # 查找占位的参数
        for args in re.findall('\%\d\$\@', text):
            excel_args = args.replace('%', '{').replace('$@', '}')
            value = text.replace(args, excel_args)
    return value
    
# 将ios国际化源文件读出的多语言写入excel文件
def write_trans_to_excel(path, trans_map):
    # 创建文件
    if not os.path.exists(path):
        with open(path, 'w+', encoding='utf-8') as file:
            file.close
    workbook = Workbook()
    sheet = workbook.active
    
    sheet.cell(row=1, column=1).value = 'key'  
    # 列数
    currentColumn = 2
    
    sort_keys = sorted(trans_map)
    for localizable_key in sort_keys:
        localizable_dict = trans_map[localizable_key]
        
        _localizable = localizable_key
        if _localizable == LanguageType.zh_hans:
            _localizable = f'简体中文:{localizable_key}'
        elif _localizable == LanguageType.zh_hant:
            _localizable = f'繁体中文:{localizable_key}'
        elif _localizable == LanguageType.en:
            _localizable = f'英文:{localizable_key}'
        elif _localizable == LanguageType.es:
            _localizable = f'西班牙语:{localizable_key}'
        elif _localizable == LanguageType.ja:
            _localizable = f'日语:{localizable_key}'
        elif _localizable == LanguageType.ko:
            _localizable = f'韩语:{localizable_key}'
        
        # 第一行描述
        localizable_desc = _localizable.replace('"', '')
        sheet.cell(row=1, column=currentColumn).value = localizable_desc
        
        currentRow = 2
        for key in localizable_dict:
            _key = key.replace('"', '')
            _string = localizable_dict[key].replace('"', '')
            # 第一列key
            sheet.cell(row=currentRow, column=1).value = _key
            # 多语言翻译
            sheet.cell(row=currentRow, column=currentColumn).value = _string
            currentRow = currentRow + 1
        
        currentColumn = currentColumn + 1

    # 设置excel样式
    for row in range(0, sheet.max_row+1):
        sheet.row_dimensions[row].height = excel_cell_height
    for col in range(1, sheet.max_column+1):
        col_name = get_column_letter(col)
        if col == 1:
            sheet.column_dimensions[col_name].width = 40
        else:
            sheet.column_dimensions[col_name].width = excel_cell_width
    
    for col in range(1, sheet.max_column+1):
        col_name = get_column_letter(col)
        for row in range(0, sheet.max_row):
            sheet[col_name][row].font = font
            sheet[col_name][row].alignment = align
    
    workbook.save(path)
    workbook.close

# 函数入口
if __name__ == '__main__':
    
    # 读取ios国际化源文件
    trans_map_ios = read_trans_file_ios()
    # 写入excel文件
    write_trans_to_excel(output_ios_excel_path, trans_map_ios)
    
    print('翻译源文件转excel成功...')

