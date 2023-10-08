import re
import os
import openpyxl
from language_type import *

# android翻译源文件文件夹
android_excel_file_path = './trans_from_excel/input/android/'

# ios
ios_excel_file_path = './trans_from_excel/input/ios/'

# web翻译源文件文件夹
web_excel_file_path = './trans_from_excel/input/web/'

# 安卓输出翻译源文件
android_output_trans_path = './trans_from_excel/output/android/'

# 安卓输出翻译源文件
ios_output_trans_path = './trans_from_excel/output/ios/'

# 安卓输出翻译源文件
web_output_trans_path = './trans_from_excel/output/web/'

# 读取excel,生成翻译映射关系 例如:{zh_hans: {'trans0001': "中文"}, en: {'trans0001': "english"}}
def read_excel(excen_path):


    # 导入多语言翻译的excel
    workbook = openpyxl.load_workbook(excen_path, data_only=True)
    # 当前工作表
    sheet = workbook.active
    # 最大行数
    max_row = sheet.max_row + 1
    # 最大列数
    max_col = sheet.max_column + 1

    # 多语言翻译的map 例如:{zh_hans: {'trans0001': "中文"}, en: {'trans0001': "english"}}
    trans_map = {}

    # 对应语言
    for langCol in range(2, max_col): 
        # 第一行语言描述 例如: 中文简体:zh_hans
        language_desc = sheet.cell(1, langCol).value
        if language_desc:
            language_desc = str(language_desc)
            # 拆分语言的类型 
            language = language_desc
            _split_language = language_desc.split(':')
            if len(_split_language) > 1:
              language = _split_language[1]
            # 翻译表根据语言初始化
            trans_lang_map = {}
            # 第二行开始 第一列的多语言key
            for row in range(2, max_row):
                excel_key = sheet.cell(row, 1).value
                if excel_key:
                    # excel中的key
                    excel_key = str(excel_key)
                    # excel中的值
                    excel_value = str(sheet.cell(row=row, column=langCol).value)
                    # key 去掉空格 大写转小写
                    output_key = excel_key.replace(' ', '').lower()
                    trans_lang_map[output_key] = excel_value
            trans_map[language] = trans_lang_map

    workbook.close()
    return trans_map

# 根据excel中的value转换拼接iOS 安卓 web格式化参数
def _convert_excel_args(text, format, platform):
    value = text
    # excel占位参数的map {占位参数: 位置}
    args_map = {}
    # 过滤{0}占位的参数
    for result in re.finditer('{[\d]}', text):
        span = result.span()
        # excel中参数的占位格式
        arg_fmt = text[span[0]:span[1]]
        # 占位参数位置
        arg_index = arg_fmt.replace('{', '').replace('}', '')
        args_map[arg_fmt] = arg_index
    for key in args_map:
        arg_index = args_map[key]
        # web 下标从0开始 excel中翻译从1开始 所以默认减1
        if platform == PlatformType.web:
          
          arg_index = int(arg_index) - 1
          arg_index = max(0, arg_index)
        # d表示格式化参数的位置 将d替换成具体位置
        arg_fmt = format.replace('d', str(arg_index))
        # 替换excel中的占位参数
        value = value.replace(key, arg_fmt)
    return value

# 逐行写入国际化源文件
def _writelines(trans_lines, dir, file_name):
    dir_path = f'{dir}'
    file_path = f'{dir_path}/{file_name}'
    # 创建文件夹
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)
    # 逐行写入源文件
    with open(file_path, 'w+', encoding='utf-8') as file:
        file.writelines(trans_lines)

# 逐行写入ios国际化源文件
def _write_ios_localizable_strings(trans_dict, name):
    dir_path = f'{ios_output_trans_path}{name}.lproj'
    trans_lines = []
    for key in trans_dict:
        #  逐行拼接源文件
        value = trans_dict[key]
        text = _convert_excel_args(value, '%d$@', PlatformType.ios)
        trans_string = f'"{key.replace(" ", "")}"="{text}";\n'
        trans_lines.append(trans_string)
    _writelines(trans_lines, dir_path, 'Localizable.strings')

# 逐行写入android国际化源文件
def _write_android_xml(trans_dict, name):
    dir_path = f'{android_output_trans_path}{name}'
    trans_lines = []
    xml_header = '<?xml version="1.0" encoding="utf-8"?>\n<resources>\n'
    trans_lines.append(xml_header)

    for key in trans_dict:
        #  逐行拼接源文件
        value = trans_dict[key]
        text = _convert_excel_args(value, '%d$s', PlatformType.android)
        trans_string = f'    <string name="{key.replace(" ", "")}">{text}</string>\n'
        trans_lines.append(trans_string)

    xml_tril = '</resources>'
    trans_lines.append(xml_tril)
    _writelines(trans_lines, dir_path, 'strings.xml')

# 逐行写入web国际化源文件
def _write_web_lang(trans_dict, name):
    dir_path = f'{web_output_trans_path}'
    trans_lines = []
    json_header = 'export default {\n'
    trans_lines.append(json_header)
    for key in trans_dict:
        #  逐行拼接源文件
        value = trans_dict[key]
        text = _convert_excel_args(value, '{d}', PlatformType.web)
        # 去掉excel中输入回车换行
        text = text.replace('\n', '')
        trans_string = f'    "{key.replace(" ", "")}": "{text}",\n'
        trans_lines.append(trans_string)

    json_tril = '}'
    trans_lines.append(json_tril)
    _writelines(trans_lines, dir_path, name)

# 将excel读出的多语言map写入android国际化源文件
def write_trans_to_file_android(trans_map):
    for key in trans_map:
        # 每种语言对应的翻译map
        trans_dict = trans_map[key]
        _write_android_xml(trans_dict, key)

# 将excel读出的多语言map写入ios国际化源文件
def write_trans_to_file_ios(trans_map):
    for key in trans_map:
        # 每种语言对应的翻译map
        trans_dict = trans_map[key]
        _write_ios_localizable_strings(trans_dict, key)

# 将excel读出的多语言map写入web国际化源文件
def write_trans_to_file_web(trans_map):
    for key in trans_map:
        # 每种语言对应的翻译map
        trans_dict = trans_map[key]
        _write_web_lang(trans_dict, f'{key}.js')

# 遍历读取file_path文件夹下的所有excel翻译文件
def filter_all_excels(file_path):
    excel_file_paths = []
    for root,dirs,files in os.walk(rf'{file_path}'): #遍历文件夹
        for f in files: #遍历文件
            if f.find('.xlsx') > 0: # 是excel文件
                dirPath = os.path.join(root, f) #拼接文件名路径
                excel_file_paths.append(dirPath)
    # print(f'excel_file_paths: {excel_file_paths}')
    return excel_file_paths

# 合并dir_path文件夹的所有excel翻译文件
def merge_all_excels(dir_path):
    all_trans_map = {}
    file_paths = filter_all_excels(dir_path)
    for path in file_paths: #遍历文件夹
        trans_map = read_excel(path)
        all_trans_map.update(trans_map)
    return all_trans_map

# 函数入口
if __name__ == '__main__':
    # 合并android目录下所有的excel翻译文件
    trans_map_android = merge_all_excels(android_excel_file_path)
    # 写入android国际化源文件
    write_trans_to_file_android(trans_map_android)

    # 合并ios目录下所有的excel翻译文件
    trans_map_ios= merge_all_excels(ios_excel_file_path)
    # 写入ios国际化源文件
    write_trans_to_file_ios(trans_map_ios)

    # 合并web目录下所有的excel翻译文件
    trans_map_web = merge_all_excels(web_excel_file_path)
    # 写入web国际化源文件
    write_trans_to_file_web(trans_map_web)
    
    print('excel转翻译源文件成功...')

